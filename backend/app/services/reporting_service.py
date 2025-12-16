from app import db, cache, mail
from app.models import User, Member, Loan, Transaction, Branch, Group, LoanProduct
from datetime import datetime, timedelta
from sqlalchemy import func, and_, or_
import pandas as pd
import json
import logging
from decimal import Decimal
import io
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from flask_mail import Message

logger = logging.getLogger(__name__)
CACHE_TIMEOUT = 300


class ReportingService:
    
    REPORT_TEMPLATES = {
        'compliance': {
            'name': 'Regulatory Compliance Report',
            'description': 'CBK and SASRA regulatory returns',
            'columns': ['total_members', 'active_members', 'total_loans', 'active_loans', 'par_ratio', 'npl_ratio', 'total_assets']
        },
        'operations': {
            'name': 'Daily Operations Summary',
            'description': 'Daily activity snapshot',
            'columns': ['date', 'new_applications', 'approvals', 'disbursements', 'repayments', 'savings_deposits']
        },
        'financial': {
            'name': 'Monthly Financial Statement',
            'description': 'Income statement and balance sheet',
            'columns': ['interest_income', 'processing_fees', 'total_expenses', 'net_profit', 'total_assets', 'total_liabilities']
        },
        'branch': {
            'name': 'Branch Performance Scorecard',
            'description': 'Branch-level KPI analysis',
            'columns': ['branch_name', 'members', 'loans_amount', 'repayment_rate', 'par_ratio', 'staff_count']
        },
        'member': {
            'name': 'Member Analytics Report',
            'description': 'Member cohort and segment analysis',
            'columns': ['member_id', 'status', 'loans_count', 'savings_balance', 'risk_score', 'lifecycle_stage']
        },
        'risk': {
            'name': 'Risk Management Report',
            'description': 'Portfolio risk and early warning analysis',
            'columns': ['risk_category', 'member_count', 'outstanding_amount', 'par_loans', 'default_probability']
        }
    }
    
    @staticmethod
    def get_available_templates():
        """Get list of all available report templates"""
        try:
            return {
                'status': 'success',
                'templates': ReportingService.REPORT_TEMPLATES,
                'count': len(ReportingService.REPORT_TEMPLATES)
            }
        except Exception as e:
            logger.error(f"Get templates error: {str(e)}")
            return {'status': 'error', 'message': str(e)}
    
    @staticmethod
    def generate_compliance_report(branch_id=None):
        """Generate regulatory compliance report"""
        try:
            members = Member.query.outerjoin(Group).all()
            if branch_id:
                members = [m for m in members if m.group and m.group.branch_id == branch_id]
            
            loans = Loan.query.all()
            if branch_id:
                loans = [l for l in loans if l.member.group and l.member.group.branch_id == branch_id]
            
            active_members = len([m for m in members if m.status == 'active'])
            active_loans = len([l for l in loans if l.status == 'active'])
            arrears_loans = len([l for l in loans if l.status in ['arrears', 'defaulted']])
            
            par_ratio = (arrears_loans / len(loans) * 100) if loans else 0
            npl_ratio = (len([l for l in loans if l.status == 'defaulted']) / len(loans) * 100) if loans else 0
            total_assets = sum(float(m.savings_account.balance) if m.savings_account else 0 for m in members)
            
            report_data = {
                'report_type': 'compliance',
                'generated_at': datetime.utcnow().isoformat(),
                'summary': {
                    'total_members': len(members),
                    'active_members': active_members,
                    'total_loans': len(loans),
                    'active_loans': active_loans,
                    'par_ratio': float(par_ratio),
                    'npl_ratio': float(npl_ratio),
                    'total_assets': float(total_assets)
                },
                'branch_id': branch_id,
                'status': 'success'
            }
            
            return report_data
        except Exception as e:
            logger.error(f"Compliance report error: {str(e)}")
            return {'status': 'error', 'message': str(e)}
    
    @staticmethod
    def generate_operations_report(days_back=1, branch_id=None):
        """Generate daily operations summary"""
        try:
            start_date = datetime.utcnow() - timedelta(days=days_back)
            
            new_apps = Loan.query.filter(Loan.application_date >= start_date).all()
            if branch_id:
                new_apps = [l for l in new_apps if l.member.group and l.member.group.branch_id == branch_id]
            
            approvals = [l for l in new_apps if l.status == 'approved']
            disbursements = [l for l in new_apps if l.status in ['active', 'completed']]
            
            transactions = Transaction.query.filter(Transaction.created_at >= start_date).all()
            if branch_id:
                transactions = [t for t in transactions if t.member.group and t.member.group.branch_id == branch_id]
            
            repayments = len([t for t in transactions if t.transaction_type == 'loan_repayment'])
            deposits = len([t for t in transactions if t.transaction_type == 'deposit'])
            
            report_data = {
                'report_type': 'operations',
                'generated_at': datetime.utcnow().isoformat(),
                'period': f'Last {days_back} day(s)',
                'summary': {
                    'new_applications': len(new_apps),
                    'approvals': len(approvals),
                    'disbursements': len(disbursements),
                    'repayments': repayments,
                    'savings_deposits': deposits,
                    'total_transactions': len(transactions)
                },
                'branch_id': branch_id,
                'status': 'success'
            }
            
            return report_data
        except Exception as e:
            logger.error(f"Operations report error: {str(e)}")
            return {'status': 'error', 'message': str(e)}
    
    @staticmethod
    def generate_financial_report(month=None, branch_id=None):
        """Generate monthly financial statement"""
        try:
            if not month:
                month = datetime.utcnow().month
            
            loans = Loan.query.all()
            if branch_id:
                loans = [l for l in loans if l.member.group and l.member.group.branch_id == branch_id]
            
            completed_loans = [l for l in loans if l.status == 'completed']
            interest_income = sum(float(l.interest_amount) for l in completed_loans)
            processing_fees = sum(float(l.charge_fee) for l in completed_loans)
            
            members = Member.query.outerjoin(Group).all()
            if branch_id:
                members = [m for m in members if m.group and m.group.branch_id == branch_id]
            
            total_assets = sum(float(m.savings_account.balance) if m.savings_account else 0 for m in members)
            total_outstanding = sum(float(l.outstanding_balance) for l in loans)
            
            report_data = {
                'report_type': 'financial',
                'generated_at': datetime.utcnow().isoformat(),
                'period': f'Month {month}',
                'summary': {
                    'interest_income': float(interest_income),
                    'processing_fees': float(processing_fees),
                    'total_revenue': float(interest_income + processing_fees),
                    'total_assets': float(total_assets),
                    'total_liabilities': float(total_outstanding),
                    'net_equity': float(total_assets - total_outstanding)
                },
                'branch_id': branch_id,
                'status': 'success'
            }
            
            return report_data
        except Exception as e:
            logger.error(f"Financial report error: {str(e)}")
            return {'status': 'error', 'message': str(e)}
    
    @staticmethod
    def generate_member_report(branch_id=None):
        """Generate member analytics report"""
        try:
            members = Member.query.outerjoin(Group).all()
            if branch_id:
                members = [m for m in members if m.group and m.group.branch_id == branch_id]
            
            member_data = []
            for member in members[:100]:
                loans = Loan.query.filter_by(member_id=member.id).all()
                savings = float(member.savings_account.balance) if member.savings_account else 0
                
                member_data.append({
                    'member_id': member.id,
                    'status': member.status,
                    'loans_count': len(loans),
                    'savings_balance': savings,
                    'risk_score': member.risk_score,
                    'lifecycle_stage': 'mature' if len(loans) > 2 else ('active' if len(loans) > 0 else 'onboarding')
                })
            
            report_data = {
                'report_type': 'member',
                'generated_at': datetime.utcnow().isoformat(),
                'total_members': len(members),
                'members': member_data,
                'branch_id': branch_id,
                'status': 'success'
            }
            
            return report_data
        except Exception as e:
            logger.error(f"Member report error: {str(e)}")
            return {'status': 'error', 'message': str(e)}
    
    @staticmethod
    def generate_risk_report(branch_id=None):
        """Generate risk management report"""
        try:
            loans = Loan.query.all()
            if branch_id:
                loans = [l for l in loans if l.member.group and l.member.group.branch_id == branch_id]
            
            by_risk = {
                'low': len([l for l in loans if l.member.risk_score < 30]),
                'medium': len([l for l in loans if 30 <= l.member.risk_score < 60]),
                'high': len([l for l in loans if 60 <= l.member.risk_score < 80]),
                'critical': len([l for l in loans if l.member.risk_score >= 80])
            }
            
            arrears_loans = [l for l in loans if l.status == 'arrears']
            defaulted_loans = [l for l in loans if l.status == 'defaulted']
            
            report_data = {
                'report_type': 'risk',
                'generated_at': datetime.utcnow().isoformat(),
                'summary': {
                    'total_loans': len(loans),
                    'low_risk': by_risk['low'],
                    'medium_risk': by_risk['medium'],
                    'high_risk': by_risk['high'],
                    'critical_risk': by_risk['critical'],
                    'arrears_loans': len(arrears_loans),
                    'defaulted_loans': len(defaulted_loans)
                },
                'risk_distribution': by_risk,
                'branch_id': branch_id,
                'status': 'success'
            }
            
            return report_data
        except Exception as e:
            logger.error(f"Risk report error: {str(e)}")
            return {'status': 'error', 'message': str(e)}
    
    @staticmethod
    def export_to_pdf(report_data, filename=None):
        """Export report to PDF format"""
        try:
            if not filename:
                filename = f"report_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.pdf"
            
            buffer = io.BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
            
            story = []
            styles = getSampleStyleSheet()
            
            title = Paragraph(f"<b>{report_data.get('report_type', 'Report')}</b>", styles['Heading1'])
            story.append(title)
            story.append(Spacer(1, 0.3*inch))
            
            generated_time = Paragraph(f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal'])
            story.append(generated_time)
            story.append(Spacer(1, 0.2*inch))
            
            summary = report_data.get('summary', {})
            data = [['Metric', 'Value']]
            for key, value in summary.items():
                if isinstance(value, float):
                    data.append([key.replace('_', ' ').title(), f"{value:,.2f}"])
                else:
                    data.append([key.replace('_', ' ').title(), str(value)])
            
            table = Table(data, colWidths=[3*inch, 2*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 14),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(table)
            doc.build(story)
            
            buffer.seek(0)
            return {'status': 'success', 'filename': filename, 'buffer': buffer}
        except Exception as e:
            logger.error(f"PDF export error: {str(e)}")
            return {'status': 'error', 'message': str(e)}
    
    @staticmethod
    def export_to_excel(report_data, filename=None):
        """Export report to Excel format"""
        try:
            if not filename:
                filename = f"report_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            wb = Workbook()
            ws = wb.active
            ws.title = report_data.get('report_type', 'Report')
            
            ws['A1'] = report_data.get('report_type', 'Report').upper()
            ws['A1'].font = Font(bold=True, size=14)
            
            ws['A3'] = 'Generated'
            ws['B3'] = datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')
            
            row = 5
            summary = report_data.get('summary', {})
            for key, value in summary.items():
                ws[f'A{row}'] = key.replace('_', ' ').title()
                ws[f'B{row}'] = value
                row += 1
            
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            return {'status': 'success', 'filename': filename, 'buffer': buffer}
        except Exception as e:
            logger.error(f"Excel export error: {str(e)}")
            return {'status': 'error', 'message': str(e)}
    
    @staticmethod
    def send_report_email(recipients, subject, report_data, attachment_buffer=None, attachment_filename=None):
        """Send report via email"""
        try:
            msg = Message(
                subject=subject,
                recipients=recipients if isinstance(recipients, list) else [recipients],
                body=f"Report generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')}\n\nReport Type: {report_data.get('report_type')}"
            )
            
            if attachment_buffer and attachment_filename:
                msg.attach(attachment_filename, 'application/octet-stream', attachment_buffer.getvalue())
            
            mail.send(msg)
            
            return {'status': 'success', 'message': 'Email sent successfully'}
        except Exception as e:
            logger.error(f"Email send error: {str(e)}")
            return {'status': 'error', 'message': str(e)}
